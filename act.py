import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook('Trani.xlsx')
sheet=wb['Sheet1']
data=sheet.cell(1,4)
data.value="updated"
for row in range(2,sheet.max_row+1):
    val=sheet.cell(row,3)
    val.value=row
    val1=sheet.cell(row,4)
    val1.value=val.value*2
values=Reference(sheet,
                 min_row=2,
                 max_row=sheet.max_row,
                 min_col=4,
                 max_col=4)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('Trani3.xlsx')

